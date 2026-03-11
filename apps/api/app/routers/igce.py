"""IGCE (Independent Government Cost Estimate) router."""
import io
import time
from datetime import datetime
from typing import Any, Dict, List, Optional
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
import uuid
import structlog

from ..dependencies import get_db
from ..models.database import IGCEProject
from ..schemas.common import PaginatedResponse
from ..tools.registry import get_registry

router = APIRouter()
logger = structlog.get_logger(__name__)


# ---------------------------------------------------------------------------
# Frontend-compatible request / response models
# ---------------------------------------------------------------------------

class FrontendLaborLine(BaseModel):
    """A single labor line as sent by the frontend form."""

    id: str
    category: str
    laborCategory: str
    year: int
    rate: float
    hours: float
    subtotal: float


class FrontendLaborCategory(BaseModel):
    """A labor category with one or more year lines."""

    id: str
    name: str
    baseRate: float
    escalationRate: float  # percentage, e.g. 2.5 means 2.5 %
    lines: List[FrontendLaborLine] = Field(default_factory=list)


class FrontendTravelEvent(BaseModel):
    """A single travel event."""

    id: str
    destination: str
    purpose: str
    duration: float       # days per trip
    frequency: float      # trips per year
    transportationCost: float
    lodging: float
    mealsAndIncidentals: float


class FrontendAssumptions(BaseModel):
    """Assumptions block from the form."""

    laborEscalation: float = 2.0
    travelCostInflation: float = 2.0
    contingency: float = 10.0
    profitMargin: float = 15.0
    notes: str = ""


class FrontendPerformancePeriod(BaseModel):
    startDate: str
    endDate: str


class FrontendIGCERequest(BaseModel):
    """Request body exactly matching the frontend IGCEInput type."""

    projectName: str
    projectDescription: str
    performancePeriod: FrontendPerformancePeriod
    laborCategories: List[FrontendLaborCategory]
    travelEvents: List[FrontendTravelEvent] = Field(default_factory=list)
    assumptions: FrontendAssumptions = Field(default_factory=FrontendAssumptions)


# ---------------------------------------------------------------------------
# Calculation helpers
# ---------------------------------------------------------------------------

def _round2(v: float) -> float:
    return round(v, 2)


def _compute_igce(req: FrontendIGCERequest) -> dict:
    """
    Pure calculation:
      1. Sum labor lines (use subtotals from frontend if present, else rate*hours).
         Apply escalation per year: cost_year_n = cost_year_1 * (1 + esc/100)^(n-1).
      2. Sum travel: annual = (transport + lodging + mie) * duration * frequency.
      3. subtotal = labor + travel
      4. contingency = subtotal * contingency% / 100
      5. profit = (subtotal + contingency) * profitMargin% / 100
      6. total = subtotal + contingency + profit
    Returns a dict matching the frontend IGCEOutput type.
    """
    assumptions = req.assumptions
    labor_by_year: Dict[int, float] = {}
    travel_by_year: Dict[int, float] = {}

    # --- Labor ---
    for lc in req.laborCategories:
        for line in lc.lines:
            year = line.year
            # Use frontend-computed subtotal when available; fallback to rate*hours
            base_cost = line.subtotal if line.subtotal > 0 else (line.rate * line.hours)
            # Apply escalation for option years (year 1 = no escalation)
            n = year - 1  # 0-based index
            escalated = base_cost * ((1 + lc.escalationRate / 100) ** n)
            labor_by_year[year] = labor_by_year.get(year, 0.0) + escalated

    labor_total = sum(labor_by_year.values())

    # If no lines were provided but categories exist, build year-1 costs from baseRate only
    if not labor_by_year and req.laborCategories:
        for lc in req.laborCategories:
            # Assume 2080 standard hours when no lines given
            cost = lc.baseRate * 2080
            labor_by_year[1] = labor_by_year.get(1, 0.0) + cost
        labor_total = sum(labor_by_year.values())

    # --- Travel ---
    for te in req.travelEvents:
        annual = (te.transportationCost + te.lodging + te.mealsAndIncidentals) * te.duration * te.frequency
        # Assign to year 1 (single-year travel; could be extended later)
        travel_by_year[1] = travel_by_year.get(1, 0.0) + annual

    travel_total = sum(travel_by_year.values())

    # --- Rollup ---
    subtotal = labor_total + travel_total
    contingency_amount = subtotal * (assumptions.contingency / 100)
    profit_amount = (subtotal + contingency_amount) * (assumptions.profitMargin / 100)
    final_total = subtotal + contingency_amount + profit_amount

    # Sensitivity (±10 % low/high)
    sensitivity = {
        "low": _round2(final_total * 0.9),
        "base": _round2(final_total),
        "high": _round2(final_total * 1.1),
    }

    # Formulas (human-readable audit trail)
    formulas: Dict[str, str] = {
        "labor_total": "sum(rate × hours) per year, escalated at lc.escalationRate% compound",
        "travel_annual": "(transportation + lodging + mie) × duration_days × frequency",
        "subtotal": "labor_total + travel_total",
        "contingency": f"subtotal × {assumptions.contingency}%",
        "profit": f"(subtotal + contingency) × {assumptions.profitMargin}%",
        "total": "subtotal + contingency + profit",
        "sensitivity_low": "total × 0.90",
        "sensitivity_high": "total × 1.10",
    }

    # Citations (BLS OEWS reference)
    citations = [
        {
            "id": "bls-oews",
            "source": "BLS Occupational Employment and Wage Statistics",
            "url": "https://www.bls.gov/oes/",
            "timestamp": datetime.utcnow().isoformat(),
            "snippet": "BLS OEWS data used as wage reference for cost estimation.",
            "relevance": 0.9,
        }
    ]

    now = datetime.utcnow().isoformat()

    return {
        "id": str(uuid.uuid4()),
        "input": {
            "projectName": req.projectName,
            "projectDescription": req.projectDescription,
            "performancePeriod": {
                "startDate": req.performancePeriod.startDate,
                "endDate": req.performancePeriod.endDate,
            },
            "laborCategories": [lc.model_dump() for lc in req.laborCategories],
            "travelEvents": [te.model_dump() for te in req.travelEvents],
            "assumptions": assumptions.model_dump(),
        },
        "summaryTotal": _round2(final_total),
        "laborTotal": _round2(labor_total),
        "travelTotal": _round2(travel_total),
        "contingencyTotal": _round2(contingency_amount),
        "profitTotal": _round2(profit_amount),
        "finalTotal": _round2(final_total),
        "laborByYear": {k: _round2(v) for k, v in sorted(labor_by_year.items())},
        "travelByYear": {k: _round2(v) for k, v in sorted(travel_by_year.items())},
        "sensitivity": sensitivity,
        "formulas": formulas,
        "citations": citations,
        "createdAt": now,
        "updatedAt": now,
    }


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.post("/calculate")
async def calculate_igce(request: FrontendIGCERequest) -> dict:
    """
    Build an IGCE from the frontend form payload.

    Accepts the frontend's IGCEInput schema (camelCase) and returns a
    response that matches the frontend's IGCEOutput type exactly so the
    IGCEResults component can render it without any mapping.
    """
    t0 = time.monotonic()
    try:
        output = _compute_igce(request)
        duration_ms = round((time.monotonic() - t0) * 1000, 1)
        logger.info(
            "igce_calculated",
            project=request.projectName,
            final_total=output["finalTotal"],
            duration_ms=duration_ms,
        )
        # Return the IGCEOutput object directly — api.ts expects `IGCEOutput`, not a wrapper
        return output
    except HTTPException:
        raise
    except Exception as e:
        logger.error("igce_calculate_failed", error=str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to calculate IGCE: {e}",
        )


@router.post("/build")
async def build_igce(request: FrontendIGCERequest) -> dict:
    """Alias for POST /igce/calculate."""
    return await calculate_igce(request)


# ---------------------------------------------------------------------------
# BLS Occupation list for keyword search
# Source: BLS OEWS May 2023 national estimates (bls.gov/oes)
# ---------------------------------------------------------------------------
_BLS_OCCUPATIONS = [
    ("11-1011", "Chief Executives", 246440, 206680),
    ("11-1021", "General and Operations Managers", 130600, 107360),
    ("11-3021", "Computer and Information Systems Managers", 176450, 169510),
    ("11-3031", "Financial Managers", 166050, 139790),
    ("11-9199", "Project Management Specialists", 98580, 94500),
    ("13-1041", "Compliance Officers", 79510, 72310),
    ("13-1071", "Human Resources Specialists", 67650, 62290),
    ("13-1111", "Management Analysts", 102270, 95290),
    ("13-1161", "Market Research Analysts", 78500, 68230),
    ("13-2011", "Accountants and Auditors", 82320, 77250),
    ("13-2051", "Financial Analysts", 101410, 96220),
    ("15-1211", "Computer Systems Analysts", 108990, 102240),
    ("15-1212", "Information Security Analysts", 120360, 112000),
    ("15-1221", "Computer and Information Research Scientists", 145080, 136620),
    ("15-1231", "Computer Network Support Specialists", 72560, 67640),
    ("15-1241", "Computer Network Architects", 131660, 126900),
    ("15-1244", "Network and Computer Systems Administrators", 92970, 90520),
    ("15-1251", "Computer Programmers", 99860, 97800),
    ("15-1252", "Software Developers", 130160, 124200),
    ("15-1253", "Software Quality Assurance Analysts and Testers", 105340, 98220),
    ("15-1254", "Web Developers", 86040, 80730),
    ("15-1255", "Web and Digital Interface Designers", 84120, 79190),
    ("15-1299", "Computer Occupations, All Other", 109020, 103690),
    ("15-2011", "Actuaries", 120970, 113990),
    ("15-2031", "Operations Research Analysts", 91040, 82360),
    ("15-2041", "Statisticians", 108490, 99960),
    ("15-2051", "Data Scientists", 124490, 108020),
    ("17-2061", "Computer Hardware Engineers", 132360, 128170),
    ("17-2072", "Electronics Engineers", 119320, 112620),
    ("17-2141", "Mechanical Engineers", 99510, 96310),
    ("17-2051", "Civil Engineers", 95110, 89940),
    ("17-2112", "Industrial Engineers", 100990, 96350),
    ("17-3023", "Electrical and Electronics Engineering Technologists", 69640, 67850),
    ("19-1042", "Medical Scientists", 105720, 80820),
    ("19-2041", "Environmental Scientists and Specialists", 78140, 73230),
    ("19-3011", "Economists", 123820, 107870),
    ("21-1022", "Healthcare Social Workers", 62840, 59870),
    ("23-1011", "Lawyers", 165170, 135740),
    ("23-2011", "Paralegals and Legal Assistants", 62290, 59200),
    ("25-1099", "Postsecondary Teachers", 89980, 80840),
    ("25-9099", "Education and Training Workers", 62440, 57350),
    ("27-3041", "Editors", 77390, 67530),
    ("27-3042", "Technical Writers", 81510, 79960),
    ("33-1021", "First-Line Supervisors of Firefighting Workers", 84600, 78790),
    ("41-4011", "Sales Representatives — Technical", 97860, 77250),
    ("43-3031", "Bookkeeping, Accounting, Auditing Clerks", 47440, 45860),
    ("43-6011", "Executive Secretaries and Executive Administrative Assistants", 63110, 60650),
    ("43-9061", "Office Clerks, General", 37210, 36010),
    ("47-2111", "Electricians", 60240, 60240),
    ("49-9071", "Maintenance and Repair Workers, General", 45580, 43640),
]


@router.get("/bls-lookup")
async def bls_lookup(q: str = "") -> dict:
    """
    Search BLS OEWS occupations by keyword.
    Returns top 5 matches with mean/median hourly rates.
    No API key required — uses built-in 2023 BLS OEWS data.
    """
    if not q.strip():
        raise HTTPException(status_code=400, detail="q parameter required")

    q_lower = q.lower()
    results = []
    for soc, title, mean_annual, median_annual in _BLS_OCCUPATIONS:
        title_lower = title.lower()
        score = 0
        if q_lower == title_lower:
            score = 20
        elif title_lower.startswith(q_lower):
            score = 12
        elif q_lower in title_lower:
            score = 8
        else:
            for word in q_lower.split():
                if len(word) > 2 and word in title_lower:
                    score += 3
        if score > 0:
            results.append({
                "soc_code": soc,
                "title": title,
                "mean_annual": mean_annual,
                "median_annual": median_annual,
                "mean_hourly": round(mean_annual / 2080, 2),
                "median_hourly": round(median_annual / 2080, 2),
                "score": score,
            })

    results.sort(key=lambda x: -x["score"])
    return {
        "results": results[:5],
        "query": q,
        "source": "BLS OEWS May 2023",
        "source_url": "https://www.bls.gov/oes/",
    }


@router.get("/perdiem-lookup")
async def perdiem_lookup(city: str = "", state: str = "") -> dict:
    """
    Get GSA per diem rates for a city/state.
    Uses live GSA API when key is configured, falls back to FY2025 published rates.
    """
    if not city.strip() or not state.strip():
        raise HTTPException(status_code=400, detail="city and state are required")

    from ..tools.gsa_perdiem import GSAPerDiemTool
    tool = GSAPerDiemTool()
    try:
        result = await tool._lookup_rates({"city": city.strip(), "state": state.strip().upper(), "year": 2025})
        return result
    except Exception as e:
        logger.error("perdiem_lookup_failed", city=city, state=state, error=str(e))
        raise HTTPException(status_code=500, detail=f"Per diem lookup failed: {e}")


@router.post("/export")
async def export_igce(result: dict) -> StreamingResponse:
    """
    Accept an IGCEOutput JSON body and return a real .xlsx file.
    The frontend POSTs the full result object here and downloads the blob.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ── helpers ──────────────────────────────────────────────────────────
        HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
        HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
        SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
        SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
        TOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")
        TOTAL_FONT = Font(bold=True, size=10)
        THIN = Side(style="thin")
        BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        def _fmt_currency(ws, cell):
            ws[cell].number_format = '$#,##0.00'

        def _header(ws, row, col, text, fill=HEADER_FILL, font=HEADER_FONT):
            c = ws.cell(row=row, column=col, value=text)
            c.fill = fill
            c.font = font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
            return c

        # ── Sheet 1: Summary ─────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Summary"
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 20

        project_name = result.get("input", {}).get("projectName", "IGCE")
        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        title_cell.value = f"IGCE – {project_name}"
        title_cell.font = Font(bold=True, size=14, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center")

        perf = result.get("input", {}).get("performancePeriod", {})
        ws["A2"] = "Performance Period"
        ws["B2"] = f"{perf.get('startDate','')[:10]}  →  {perf.get('endDate','')[:10]}"
        ws["A2"].font = Font(italic=True)

        ws.append([])

        _header(ws, 4, 1, "Cost Element")
        _header(ws, 4, 2, "Amount ($)")

        rows = [
            ("Labor Total",       result.get("laborTotal", 0)),
            ("Travel Total",      result.get("travelTotal", 0)),
            ("Subtotal",          result.get("laborTotal", 0) + result.get("travelTotal", 0)),
            ("Contingency",       result.get("contingencyTotal", 0)),
            ("Profit",            result.get("profitTotal", 0)),
        ]
        for i, (label, val) in enumerate(rows, start=5):
            ws.cell(row=i, column=1, value=label).border = BORDER
            c = ws.cell(row=i, column=2, value=val)
            c.number_format = '$#,##0.00'
            c.border = BORDER

        # Final total row
        total_row = 5 + len(rows)
        tc = ws.cell(row=total_row, column=1, value="TOTAL ESTIMATE")
        tc.fill = TOTAL_FILL
        tc.font = TOTAL_FONT
        tc.border = BORDER
        vc = ws.cell(row=total_row, column=2, value=result.get("finalTotal", 0))
        vc.number_format = '$#,##0.00'
        vc.fill = TOTAL_FILL
        vc.font = TOTAL_FONT
        vc.border = BORDER

        # Sensitivity
        ws.append([])
        sens_row = total_row + 2
        _header(ws, sens_row, 1, "Sensitivity Analysis")
        _header(ws, sens_row, 2, "Value ($)")
        sensitivity = result.get("sensitivity", {})
        for j, (key, label) in enumerate([("low", "Low (−10%)"), ("base", "Base"), ("high", "High (+10%)")], 1):
            r = sens_row + j
            ws.cell(row=r, column=1, value=label).border = BORDER
            c = ws.cell(row=r, column=2, value=sensitivity.get(key, 0))
            c.number_format = '$#,##0.00'
            c.border = BORDER

        # ── Sheet 2: Labor Breakdown ─────────────────────────────────────────
        ws2 = wb.create_sheet("Labor Breakdown")
        ws2.column_dimensions["A"].width = 12
        ws2.column_dimensions["B"].width = 20
        ws2.column_dimensions["C"].width = 28
        ws2.column_dimensions["D"].width = 14
        ws2.column_dimensions["E"].width = 14
        ws2.column_dimensions["F"].width = 14

        _header(ws2, 1, 1, "Year", SUBHEADER_FILL, SUBHEADER_FONT)
        _header(ws2, 1, 2, "Category", SUBHEADER_FILL, SUBHEADER_FONT)
        _header(ws2, 1, 3, "Labor Category", SUBHEADER_FILL, SUBHEADER_FONT)
        _header(ws2, 1, 4, "Rate ($/hr)", SUBHEADER_FILL, SUBHEADER_FONT)
        _header(ws2, 1, 5, "Hours", SUBHEADER_FILL, SUBHEADER_FONT)
        _header(ws2, 1, 6, "Subtotal ($)", SUBHEADER_FILL, SUBHEADER_FONT)

        lc_row = 2
        for lc in result.get("input", {}).get("laborCategories", []):
            for line in lc.get("lines", []):
                ws2.cell(row=lc_row, column=1, value=line.get("year")).border = BORDER
                ws2.cell(row=lc_row, column=2, value=lc.get("name")).border = BORDER
                ws2.cell(row=lc_row, column=3, value=line.get("laborCategory")).border = BORDER
                rc = ws2.cell(row=lc_row, column=4, value=line.get("rate", 0))
                rc.number_format = '$#,##0.00'
                rc.border = BORDER
                ws2.cell(row=lc_row, column=5, value=line.get("hours", 0)).border = BORDER
                sc = ws2.cell(row=lc_row, column=6, value=line.get("subtotal", 0))
                sc.number_format = '$#,##0.00'
                sc.border = BORDER
                lc_row += 1

        # Labor by year totals
        if lc_row > 2:
            ws2.append([])
            lc_row += 1
            _header(ws2, lc_row, 1, "Year", SUBHEADER_FILL, SUBHEADER_FONT)
            _header(ws2, lc_row, 6, "Year Total ($)", SUBHEADER_FILL, SUBHEADER_FONT)
            for year, amt in sorted(result.get("laborByYear", {}).items()):
                lc_row += 1
                ws2.cell(row=lc_row, column=1, value=int(year)).border = BORDER
                c = ws2.cell(row=lc_row, column=6, value=amt)
                c.number_format = '$#,##0.00'
                c.border = BORDER

        # ── Sheet 3: Assumptions ─────────────────────────────────────────────
        ws3 = wb.create_sheet("Assumptions")
        ws3.column_dimensions["A"].width = 28
        ws3.column_dimensions["B"].width = 20

        _header(ws3, 1, 1, "Assumption", SUBHEADER_FILL, SUBHEADER_FONT)
        _header(ws3, 1, 2, "Value", SUBHEADER_FILL, SUBHEADER_FONT)

        assumptions = result.get("input", {}).get("assumptions", {})
        assumption_rows = [
            ("Labor Escalation (%)", assumptions.get("laborEscalation", "")),
            ("Travel Cost Inflation (%)", assumptions.get("travelCostInflation", "")),
            ("Contingency (%)", assumptions.get("contingency", "")),
            ("Profit Margin (%)", assumptions.get("profitMargin", "")),
            ("Notes", assumptions.get("notes", "")),
        ]
        for i, (label, val) in enumerate(assumption_rows, start=2):
            ws3.cell(row=i, column=1, value=label).border = BORDER
            ws3.cell(row=i, column=2, value=val).border = BORDER

        # Formulas audit trail
        ws3.append([])
        frow = 2 + len(assumption_rows) + 1
        _header(ws3, frow, 1, "Formula", SUBHEADER_FILL, SUBHEADER_FONT)
        _header(ws3, frow, 2, "Description", SUBHEADER_FILL, SUBHEADER_FONT)
        for k, v in result.get("formulas", {}).items():
            frow += 1
            ws3.cell(row=frow, column=1, value=k).border = BORDER
            ws3.cell(row=frow, column=2, value=v).border = BORDER

        # ── Stream the workbook ───────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in project_name)
        filename = f"IGCE_{safe_name}.xlsx"

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except Exception as e:
        logger.error("igce_export_failed", error=str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export IGCE: {e}",
        )


class IGCEEstimateRequest(BaseModel):
    """IGCE estimate request."""

    scope: str
    labor_hours: float
    labor_rate: float
    materials_cost: float
    overhead_percentage: float
    profit_margin_percentage: float


class IGCEEstimateResponse(BaseModel):
    """IGCE estimate response."""

    total_labor: float
    total_materials: float
    subtotal: float
    overhead: float
    profit: float
    total_estimate: float


class IGCEProjectCreate(BaseModel):
    """Create IGCE project request."""

    title: str
    assumptions: dict


class IGCEProjectResponse(BaseModel):
    """IGCE project response."""

    id: uuid.UUID
    title: str
    assumptions: dict
    result: Optional[dict]
    created_at: datetime
    updated_at: datetime

    class Config:
        json_encoders = {datetime: lambda v: v.isoformat()}


@router.post("/estimate", response_model=IGCEEstimateResponse)
async def estimate_igce(request: IGCEEstimateRequest) -> IGCEEstimateResponse:
    """
    Run IGCE computation.

    Args:
        request: IGCE estimate parameters

    Returns:
        Cost estimate
    """
    try:
        # Calculate costs
        total_labor = request.labor_hours * request.labor_rate
        total_materials = request.materials_cost
        subtotal = total_labor + total_materials

        overhead = subtotal * (request.overhead_percentage / 100)
        adjusted_subtotal = subtotal + overhead

        profit = adjusted_subtotal * (request.profit_margin_percentage / 100)
        total_estimate = adjusted_subtotal + profit

        logger.info(
            "igce_estimate_calculated",
            scope=request.scope,
            total_estimate=total_estimate,
        )

        return IGCEEstimateResponse(
            total_labor=round(total_labor, 2),
            total_materials=round(total_materials, 2),
            subtotal=round(subtotal, 2),
            overhead=round(overhead, 2),
            profit=round(profit, 2),
            total_estimate=round(total_estimate, 2),
        )
    except Exception as e:
        logger.error("igce_estimate_failed", error=str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to calculate estimate",
        )


@router.get("/projects", response_model=PaginatedResponse[IGCEProjectResponse])
async def list_igce_projects(
    skip: int = 0,
    limit: int = 20,
    db: AsyncSession = Depends(get_db),
) -> PaginatedResponse[IGCEProjectResponse]:
    """
    List saved IGCE projects.

    Args:
        skip: Pagination offset
        limit: Results per page
        db: Database session

    Returns:
        Paginated list of projects
    """
    try:
        # TODO: Filter by authenticated user
        result = await db.execute(
            select(IGCEProject)
            .order_by(IGCEProject.updated_at.desc())
            .offset(skip)
            .limit(limit)
        )
        projects = result.scalars().all()

        count_result = await db.execute(select(IGCEProject))
        total = len(count_result.scalars().all())

        items = [
            IGCEProjectResponse(
                id=p.id,
                title=p.title,
                assumptions=p.assumptions_json,
                result=p.result_json,
                created_at=p.created_at,
                updated_at=p.updated_at,
            )
            for p in projects
        ]

        pages = (total + limit - 1) // limit
        return PaginatedResponse(
            items=items,
            total=total,
            skip=skip,
            limit=limit,
            pages=pages,
        )
    except Exception as e:
        logger.error("list_igce_projects_failed", error=str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to list projects",
        )


@router.post("/projects", response_model=IGCEProjectResponse)
async def create_igce_project(
    request: IGCEProjectCreate,
    db: AsyncSession = Depends(get_db),
) -> IGCEProjectResponse:
    """
    Save a new IGCE project.

    Args:
        request: Project creation request
        db: Database session

    Returns:
        Created project
    """
    try:
        project = IGCEProject(
            id=uuid.uuid4(),
            user_id=uuid.uuid4(),  # TODO: Get from auth
            title=request.title,
            assumptions_json=request.assumptions,
        )
        db.add(project)
        await db.commit()

        logger.info("igce_project_created", project_id=str(project.id))

        return IGCEProjectResponse(
            id=project.id,
            title=project.title,
            assumptions=project.assumptions_json,
            result=project.result_json,
            created_at=project.created_at,
            updated_at=project.updated_at,
        )
    except Exception as e:
        logger.error("create_igce_project_failed", error=str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create project",
        )


@router.get("/projects/{project_id}", response_model=IGCEProjectResponse)
async def get_igce_project(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
) -> IGCEProjectResponse:
    """
    Get IGCE project details.

    Args:
        project_id: Project ID
        db: Database session

    Returns:
        Project details
    """
    try:
        result = await db.execute(
            select(IGCEProject).where(IGCEProject.id == project_id)
        )
        project = result.scalars().first()

        if not project:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Project not found",
            )

        return IGCEProjectResponse(
            id=project.id,
            title=project.title,
            assumptions=project.assumptions_json,
            result=project.result_json,
            created_at=project.created_at,
            updated_at=project.updated_at,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error("get_igce_project_failed", error=str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to get project",
        )


@router.put("/projects/{project_id}", response_model=IGCEProjectResponse)
async def update_igce_project(
    project_id: uuid.UUID,
    request: IGCEProjectCreate,
    db: AsyncSession = Depends(get_db),
) -> IGCEProjectResponse:
    """
    Update IGCE project.

    Args:
        project_id: Project ID
        request: Update request
        db: Database session

    Returns:
        Updated project
    """
    try:
        result = await db.execute(
            select(IGCEProject).where(IGCEProject.id == project_id)
        )
        project = result.scalars().first()

        if not project:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Project not found",
            )

        project.title = request.title
        project.assumptions_json = request.assumptions
        await db.commit()

        logger.info("igce_project_updated", project_id=str(project_id))

        return IGCEProjectResponse(
            id=project.id,
            title=project.title,
            assumptions=project.assumptions_json,
            result=project.result_json,
            created_at=project.created_at,
            updated_at=project.updated_at,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error("update_igce_project_failed", error=str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to update project",
        )
